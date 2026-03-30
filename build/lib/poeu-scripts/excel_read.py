
# Functions for reading off Excel sheets and returning their data.
# I wanna keep everything relating to Excel IO in here because the goal is to eventually use this 
#   program from a game interface. Returns and stuff should be as generic as possible, OpenPyXL 
#   should ONLY be used in this file.


import openpyxl as opxl
import os
import string
from powerplant import PowerPlant


temp_excel_path = 'git_ignore\\CE4321_GridOptimizer_v3_OLD.xlsx'  # this is just for testing, an actual run will take path from caller


# gets plant info from excel sheet and stores it in a list of PowerPlant objects
def get_plants(path: string = temp_excel_path):
    try:  # if all error checks pass, it'll continue on to execution.
        if os.path.exists == False:
            raise FileNotFoundError('File does not exist. Check the provided path and try again.')
        elif os.path.splitext(path)[1] != '.xlsx':
            raise TypeError('Filetype not xlsx.')  # note that opxl can't process xls files without another package
        else:  # all checks passed, assume we're working with a valid excel sheet from here

            excel_sheet  = opxl.load_workbook(path).active  # assumes the current active sheet is the one with the info
            rows = excel_sheet.iter_rows(min_row=7, max_row=22, values_only=True)  # iterates over only the rows with plant info

            plants: list[PowerPlant] = []

            for row in rows:
                # index of cell is (actual column number - 1)
                # or you could just like. start counting the columns from zero. rather than one. lmao
                if row[0] is not None:
                    id = row[0]
                    name = row[1]
                    type = row[2]
                    demand = row[11]

                    # these are dependent on amount of sun and wind
                    min_output = row[6]
                    max_output = row[8]

                    conditions = get_conditions(path)  # they call me john programming

                    if str(min_output)[0] == '=': 
                        split_string = min_output.split(',')
                        if type.lower() == "wind":
                            min_output = conditions[2] * float(split_string[1])  # multiply by wind constraint
                        elif type.lower == "solar":
                            min_output = conditions[1] * float(split_string[1])  # multiply by sun constraint
                        else: print(f'Error converting from Excel IF.')

                    if str(max_output)[0] == '=': 
                        split_string = max_output.split(',')
                        if type.lower() == "wind":
                            max_output = conditions[2] * float(split_string[1])  # multiply by wind constraint
                        elif type.lower == "solar":
                            max_output = conditions[1] * float(split_string[1])  # multiply by sun constraint
                        else: print(f'Error converting from Excel IF.')

                    # these guys are sometimes excel =if formulas, so process a bit before adding them to the plant
                    plant_cost = row[4]
                    wholesale_cost = row[13]

                    # if plant cost is an if statement, it'll be in the form =IF(demand>a certain amount, 0.11, else 0.12)
                    if str(plant_cost)[0] == '=':  
                        split_string = plant_cost.split(', ')
                        price_cutoff = float((split_string[0].split('>'))[1])
                        cost_if_true = float(split_string[1])
                        cost_if_false = float((split_string[2].split(')'))[0])

                        if demand > price_cutoff: plant_cost = cost_if_true
                        else: plant_cost = cost_if_false

                    # same deal as above but for wholesale costs
                    # =IF(demand>0, plant cost, else 0)
                    if str(wholesale_cost)[0] == '=':
                        if demand > 0: wholesale_cost = plant_cost
                        else: wholesale_cost = 0.0

                    plants.append(PowerPlant(id, name, type, plant_cost, min_output, max_output, demand, wholesale_cost))

            return plants

    except Exception as e:  # catches all errors
        print(f'An error occurred:\n    {e}')


# gets current conditions from excel sheet
# yeah parts are copied from get_plants. i cooked on that are you kidding of course im gonna reuse it
def get_conditions(path: string = temp_excel_path):
    try:  # if all error checks pass, it'll continue on to execution.
        if os.path.exists == False:
            raise FileNotFoundError('File does not exist. Check the provided path and try again.')
        elif os.path.splitext(path)[1] != '.xlsx':
            raise TypeError('Filetype not xlsx.')  # note that opxl can't process xls files without another package
        else:  # all checks passed, assume we're working with a valid excel sheet from here

            excel_sheet  = opxl.load_workbook(path).active  # assumes the current active sheet is the one with the info
            rows = excel_sheet.iter_rows(min_row=24, max_row=24, values_only=True)  # iterates over only the rows with condition info

            conditions: list[float] = []

            for row in rows:
                conditions.append(row[4])  # total power demand
                conditions.append(float(row[6]))  # amount of sun (float between 0-1)
                conditions.append(float(row[8])) # amount of wind (float between 0-1)

            return conditions

    except Exception as e:  # catches all errors
            print(f'An error occurred:\n    {e}')


# TESTING AREA. STUFF BELOW HERE WILL BE DELETED EVENTUALLY.
# plants = get_plants(temp_excel_path)
# for p in plants: 
#     print(f'''{p.id},'{p.name}', '{p.type}', {p.plant_cost}, {p.min_output}, {p.max_output}, {p.demand}, {p.wholesale_cost}''')
